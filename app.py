from flask import Flask, request, send_file, render_template
import pandas as pd
import io
import os
import openpyxl
import random

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process():
    file1 = request.files.get('file1')
    file2 = request.files.get('file2')
    if file1 and file1.filename != '':
        return process_csv(file1)
    elif file2 and file2.filename != '':
        return process_csv_v2(file2)
    else:
        return 'No selected file', 400

def process_csv(file):
    if file.filename == '':
        return 'No selected file', 400
    if file:
        file_ext = os.path.splitext(file.filename)[1]
        if file_ext == ".csv":
            df = pd.read_csv(file)
            df = process_dataframe(df)
            output = io.StringIO()
            df.to_csv(output, index=False)
            output.seek(0)
            return send_file(output, mimetype='text/csv', as_attachment=True, attachment_filename='output.csv')
        elif file_ext == ".xlsx":
            file.save("temp.xlsx")
            process_xlsx("temp.xlsx")
            return send_file("temp.xlsx", as_attachment=True, download_name="processed_file.xlsx")
        else:
            return 'Unsupported file type', 400

def process_dataframe(df):
    for index, row in df.iterrows():
        target_sum = row['G']
        if not isinstance(target_sum, int) or target_sum < 0 or target_sum > 5:
            continue
        combination = generate_combinations(target_sum)
        if combination:
            df.at[index, 'C'], df.at[index, 'D'], df.at[index, 'E'], df.at[index, 'F'] = combination
    return df

def generate_combinations(target_sum):
    valid_combinations = []
    for c in [0, 1]:
        for d in [0, 1]:
            for e in [0, 1, 2]:
                for f in [0, 1]:
                    if c + d + e + f == target_sum:
                        valid_combinations.append((c, d, e, f))
    if valid_combinations:
        return random.choice(valid_combinations)
    return None

def process_xlsx(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        target_sum = ws[f'G{row}'].value
        if not isinstance(target_sum, int) or target_sum < 0 or target_sum > 5:
            continue
        combination = generate_combinations(target_sum)
        if combination:
            ws[f'C{row}'], ws[f'D{row}'], ws[f'E{row}'], ws[f'F{row}'] = combination
    wb.save(filename)

def process_csv_v2(file):
    if file.filename == '':
        return 'No selected file', 400
    if file:
        file_ext = os.path.splitext(file.filename)[1]
        if file_ext == ".csv":
            df = pd.read_csv(file)
            df = process_dataframe_v2(df)
            output = io.StringIO()
            df.to_csv(output, index=False)
            output.seek(0)
            return send_file(output, mimetype='text/csv', as_attachment=True, attachment_filename='output_v2.csv')
        elif file_ext == ".xlsx":
            file.save("temp_v2.xlsx")
            process_xlsx_v2("temp_v2.xlsx")
            return send_file("temp_v2.xlsx", as_attachment=True, download_name="processed_file_v2.xlsx")
        else:
            return 'Unsupported file type', 400

def process_dataframe_v2(df):
    for index, row in df.iterrows():
        target_sum = row['I']
        combination = generate_combinations_v2(target_sum)
        if combination:
            df.at[index, 'C'], df.at[index, 'D'], df.at[index, 'E'], df.at[index, 'F'], df.at[index, 'G'], df.at[index, 'H'] = combination
    return df

def generate_combinations_v2(target_sum):
    valid_combinations = []
    for c in [round(i*0.1, 2) for i in range(9)]:
        for d in [round(i*0.1, 2) for i in range(9)]:
            for e in [round(i*0.1, 2) for i in range(9)]:
                for f in [round(i*0.1, 2) for i in range(9)]:
                    for g in [round(i*0.1, 2) for i in range(33)]:
                        for h in [round(i*0.1, 2) for i in range(17)]:
                            if round(c + d + e + f + g + h, 2) == target_sum:
                                valid_combinations.append((c, d, e, f, g, h))
    if valid_combinations:
        return random.choice(valid_combinations)
    return None

def process_xlsx_v2(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        target_sum = ws[f'I{row}'].value
        combination = generate_combinations_v2(target_sum)
        if combination:
            ws[f'C{row}'], ws[f'D{row}'], ws[f'E{row}'], ws[f'F{row}'], ws[f'G{row}'], ws[f'H{row}'] = combination
    wb.save(filename)

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0')
